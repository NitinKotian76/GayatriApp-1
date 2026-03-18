from django.conf import settings
from django.contrib.messages.views import SuccessMessageMixin
from django.views.generic import FormView
from django.contrib import messages
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.urls import (reverse_lazy, reverse)
from django.views import View

from ...form_files import (helperFunct as hf, millsoftForm as mf)
from ...models import TProduction, MItem, MCustomer, MAgent, MLocation

from weasyprint import HTML
from io import BytesIO
import shutil
import os
import uuid
from openpyxl import Workbook, load_workbook
import logging

logger = logging.getLogger(__name__)

class RChallanCreateView(View):
    """Create a challan PDF and redirect to its download URL."""

    def post(self, request, *args, **kwargs):
        logger.debug("the pdf view ran")
        pdf_buffer = BytesIO()
        template = request.POST.get("template")
        context = request.POST.get("context")
        template_path = os.path.join(
            settings.MEDIA_ROOT, "ReportTemplates", "DemoTemplate.html")
        html_string = render_to_string(
            "DemoTemplate.html", context
        )
        logger.debug(html_string)
        HTML(string=html_string).write_pdf(pdf_buffer)

        filename = f"challan_{uuid.uuid4().hex[:8]}.pdf"
        pdf_path = os.path.join(settings.MEDIA_ROOT, "Challans", filename)
        os.makedirs(os.path.dirname(pdf_path), exist_ok=True)

        with open(pdf_path, "wb") as f:
            f.write(pdf_buffer.getvalue())

        download_url = reverse("invoice:download_challan", args=[filename])

        messages.success(request, "pdf created")
        response = HttpResponse("pdf created")
        response["HX-Redirect"] = download_url
        return response




class RInvoiceCreateView(View):
    """Create a invoice PDF and redirect to its download URL."""

    def post(self, request, *args, **kwargs):
        return HttpResponse(status=204)

    def get(self, request, *args, **kwargs):
        return HttpResponse(status=204)


class RDispatchDetailsCreateView(View):
    """Create a dispatch details PDF and redirect to its download URL."""

    def post(self, request, *args, **kwargs):
        return HttpResponse(status=204)

    def get(self, request, *args, **kwargs):
        return HttpResponse(status=204)


class RGatePassCreateView(View):
    """Create a gate pass PDF and redirect to its download URL."""

    def post(self, request, *args, **kwargs):
        return HttpResponse(status=204)

    def get(self, request, *args, **kwargs):
        return HttpResponse(status=204)


class RStockCreateView(SuccessMessageMixin, FormView):
    """Create a stock PDF and redirect to its download URL."""
    template_name = "partials/forms.html"
    form_class = mf.RStockForm
    success_url = reverse_lazy("invoice:RStockCreateView")
    success_message = "stock report created"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context["buttons"] = hf.button(
            value="submit",
            hx_req_type="hx-post",
            hx_req=reverse_lazy("invoice:RStockCreateView")
        )
        return context

    def post(self, request, *args, **kwargs):
        date = request.POST.get("date")
        type_of_report = request.POST.get("type_of_report")
        if type_of_report == "1":
            logger.debug("all")
            data = TProduction.objects.filter(stk=True,rdate__lte=date) 
        elif type_of_report == "2":
            logger.debug("group by category")
            data = TProduction.objects.filter(stk=True,rdate__lte=date).values("category")
        data = data.values()

        logger.debug(f"data: {data}")
        report(data)
        logger.debug("report method called")
        messages.success(request, "stock report created")
        response = HttpResponse(status=204)
        response["HX-Redirect"] = reverse("invoice:DownloadExcelView", args=["stock_report.xlsx"])
        return response



#TODO: move this to a separate file and rewrite to use placeholder tags rather than hardcoded values

def report(data: list[dict]):    
    logger.debug(f"data: {data}")
    
    template_file_path = os.path.join(
        settings.MEDIA_ROOT, "Reports", "Templates", "Stock.xlsx")
    logger.debug(f"template file path: {template_file_path}")
    if not os.path.exists(template_file_path):
        return HttpResponse("template file not found", status=404)
    shutil.copyfile(template_file_path, os.path.join(
        settings.MEDIA_ROOT, "Reports", "stock_report.xlsx"))

    report_file_path = os.path.join(
        settings.MEDIA_ROOT, "Reports", "stock_report.xlsx")
    logger.debug(f"report file path: {report_file_path}")

        
    wb = load_workbook(filename=template_file_path)
    ws = wb.active
    data_len = len(data)
    # Unmerge any merged cells that overlap the data area so we can write to every cell
    # for merged_range in list(ws.merged_cells.ranges):
    #     if (
    #         merged_range.min_row <= data_len + 1
    #         and merged_range.max_row >= 1
    #         and merged_range.min_col <= 20
    #         and merged_range.max_col >= 5
    #     ):
    #         ws.unmerge_cells(str(merged_range))
    for i in range(1, data_len + 1):
        row = data[i - 1]
        rowno = i + 5
        ws.cell(row=rowno, column=1, value=row["rdate"].__str__())
        ws.cell(row=rowno, column=2, value=MItem.objects.get(pk=row["itemcode_id"]).itemcode)
        ws.cell(row=rowno, column=3, value=row["type_of_reel_sheet"])# local or export
        ws.cell(row=rowno, column=4, value=row["type_of_reel_sheet"])
        ws.cell(row=rowno, column=5, value=MItem.objects.get(pk=row["itemcode_id"]).size)
        ws.cell(row=rowno, column=6, value="x")
        ws.cell(row=rowno, column=7, value=row["length"])
        ws.cell(row=rowno, column=8, value=MItem.objects.get(pk=row["itemcode_id"]).gsm)
        ws.cell(row=rowno, column=9, value=row["noofsheet"])
        ws.cell(row=rowno, column=10, value=row["reamwt"])
        ws.cell(row=rowno, column=11, value=row["noofream"])
        ws.cell(row=rowno, column=12, value=row["noofbdls"])
        ws.cell(row=rowno, column=13, value=row["excise_from"])
        ws.cell(row=rowno, column=14, value=row["excise_to"])
        ws.cell(row=rowno, column=15, value=row["rate"])
        ws.cell(row=rowno, column=16, value=MCustomer.objects.get(pk=row["custid_id"]).custname)
        ws.cell(row=rowno, column=17, value=MAgent.objects.get(pk=row["agentid_id"]).agentname)
        ws.cell(row=rowno, column=18, value=MLocation.objects.get(pk=row["locationid_id"]).location)
        ws.cell(row=rowno, column=19, value=row["indentno"])
        ws.cell(row=rowno, column=20, value=row["lotno"])
    wb.save(report_file_path)

    return report_file_path


class RPendingOrderCreateView(View):
    """Create a pending order PDF and redirect to its download URL."""
    template_name = "partials/forms.html"
    form_class = mf.RStockForm
    success_url = reverse_lazy("invoice:RStockCreateView")
    success_message = "stock report created"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context["buttons"] = hf.button(
            value="submit",
            hx_req_type="hx-post",
            hx_req=reverse_lazy("invoice:RStockCreateView")
        )
        return context

    def post(self, request, *args, **kwargs):
        date = request.POST.get("date")
        type_of_report = request.POST.get("type_of_report")
        if type_of_report == "1":
            data = TProduction.objects.filter(stk=True,rdate__lte=date) 
        elif type_of_report == "2":
            data = TProduction.objects.filter(stk=True,rdate__lte=date).values("category")
        data = data.values()

        logger.debug(f"data: {data}")
        report(data)
        logger.debug("report method called")
        messages.success(request, "stock report created")
        response = HttpResponse(status=204)
        response["HX-Redirect"] = reverse("invoice:DownloadExcelView", args=["stock_report.xlsx"])
        return response



class RProdRecordCreateView(View):
    """Create a production record PDF and redirect to its download URL."""

    def post(self, request, *args, **kwargs):
        return HttpResponse(status=204)


class RDispatchDetailsCreateView(SuccessMessageMixin, FormView):
    template_name = "partials/forms.html"
    form_class = mf.RDispatchForm
